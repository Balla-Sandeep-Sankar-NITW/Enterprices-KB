"""
Document Processing Service
- Text extraction: PyMuPDF, python-docx, openpyxl (pure Python, no ML)
- Chunking: fixed-size word-based with overlap
- No local models, no GPU
"""
import os
import uuid
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException, status

from app.core.config import settings
from app.models.models import Document, DocumentStatus


# ─────────────────────────────────────────
# File Storage
# ─────────────────────────────────────────

def get_upload_path(department_id: int) -> Path:
    path = Path(settings.UPLOAD_DIR) / str(department_id)
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_uploaded_file(file: UploadFile, department_id: int) -> Dict[str, Any]:
    ext = Path(file.filename).suffix.lower().lstrip(".")
    if ext not in settings.allowed_extensions_list:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File type '{ext}' not allowed. Allowed: {', '.join(settings.allowed_extensions_list)}"
        )

    content = file.file.read()
    if len(content) > settings.max_file_size_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File size exceeds {settings.MAX_FILE_SIZE_MB}MB limit"
        )

    stored_filename = f"{uuid.uuid4().hex}.{ext}"
    upload_dir = get_upload_path(department_id)
    file_path = upload_dir / stored_filename

    with open(file_path, "wb") as f:
        f.write(content)

    return {
        "original_filename": file.filename,
        "stored_filename": stored_filename,
        "file_path": str(file_path),
        "file_size": len(content),
        "file_type": ext,
        "mime_type": file.content_type,
    }


def delete_file(file_path: str):
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        pass


# ─────────────────────────────────────────
# Text Extraction
# ─────────────────────────────────────────

def extract_pdf(file_path: str) -> List[Dict[str, Any]]:
    import fitz
    pages = []
    doc = fitz.open(file_path)
    for i in range(len(doc)):
        text = doc[i].get_text("text").strip()
        if text:
            pages.append({"page": i + 1, "text": text})
    doc.close()
    return pages


def extract_docx(file_path: str) -> List[Dict[str, Any]]:
    from docx import Document as DocxDoc
    doc = DocxDoc(file_path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [{"page": 1, "text": text}] if text else []


def extract_txt(file_path: str) -> List[Dict[str, Any]]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read().strip()
    return [{"page": 1, "text": text}] if text else []


def extract_xlsx(file_path: str) -> List[Dict[str, Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    pages = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(c) for c in row if c is not None)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            pages.append({"page": len(pages) + 1, "text": f"[Sheet: {sheet_name}]\n" + "\n".join(rows)})
    wb.close()
    return pages


def extract_text(file_path: str, file_type: str) -> List[Dict[str, Any]]:
    extractors = {"pdf": extract_pdf, "docx": extract_docx, "txt": extract_txt, "xlsx": extract_xlsx}
    fn = extractors.get(file_type.lower())
    if not fn:
        raise ValueError(f"No extractor for type: {file_type}")
    return fn(file_path)


# ─────────────────────────────────────────
# Chunking
# ─────────────────────────────────────────

def chunk_text(pages: List[Dict], chunk_size: int = 500, overlap: int = 50) -> List[Dict]:
    chunks = []
    idx = 0
    for page_data in pages:
        words = page_data["text"].split()
        if not words:
            continue
        start = 0
        while start < len(words):
            end = min(start + chunk_size, len(words))
            text = " ".join(words[start:end])
            if text.strip():
                chunks.append({
                    "text": text,
                    "page": page_data["page"],
                    "chunk_index": idx,
                })
                idx += 1
            start += chunk_size - overlap
    return chunks


# ─────────────────────────────────────────
# Full Processing Pipeline
# ─────────────────────────────────────────

async def process_document(db: Session, document_id: int):
    """
    Full async pipeline:
    1. Extract text from file
    2. Chunk text
    3. Embed via Voyage AI (API call)
    4. Store in Pinecone (API call)
    5. Update DB record
    """
    from app.services.vector_store import store_chunks

    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        return

    try:
        doc.status = DocumentStatus.PROCESSING
        db.commit()

        # 1. Extract
        pages = extract_text(doc.file_path, doc.file_type)
        if not pages:
            raise ValueError("No text extracted from document")

        # 2. Chunk
        chunks = chunk_text(pages, chunk_size=500, overlap=50)
        if not chunks:
            raise ValueError("Document produced no chunks after chunking")

        # 3. Add metadata
        for chunk in chunks:
            chunk["metadata"] = {
                "document_id": doc.id,
                "document_title": doc.title,
                "department_id": doc.department_id,
                "file_type": doc.file_type,
                "page": chunk["page"],
            }

        # 4. Embed + store in Pinecone (async API calls)
        await store_chunks(
            department_id=doc.department_id,
            document_id=doc.id,
            chunks=chunks
        )

        # 5. Update DB
        doc.status = DocumentStatus.PROCESSED
        doc.total_pages = max(c["page"] for c in chunks)
        doc.total_chunks = len(chunks)
        doc.processed_at = datetime.utcnow()
        db.commit()
        print(f"[Processing] ✅ Document {document_id} done: {len(chunks)} chunks")

    except Exception as e:
        print(f"[Processing] ❌ Document {document_id} failed: {e}")
        doc.status = DocumentStatus.FAILED
        doc.processing_error = str(e)
        db.commit()
